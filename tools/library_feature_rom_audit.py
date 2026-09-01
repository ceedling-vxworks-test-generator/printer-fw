#!/usr/bin/env python3
"""library_feature_rom_audit.py — C++ standard library feature usage /
rough ROM (code size) impact survey.

Recursively scans .hpp/.cpp files under the given directories, finds
every class/struct definition (reusing the lightweight brace-based
parser from coverage_audit.py), and reports which C++ standard-library
features each class actually uses (std::vector, std::string,
exceptions, RTTI, iostream, std::thread, virtual functions, ...),
alongside a rough per-feature ROM (flash/code size) impact estimate.

IMPORTANT — the ROM numbers are *rules-of-thumb*, not measurements.
They come from general embedded-C++ experience (see the "目安KB"
column in the reference table this script writes to Sheet 1), and
real cost depends heavily on the compiler, optimization flags, linker
dead-code elimination, and target architecture. Treat the per-class
"推定ROM影響" total as a *relative* ranking tool ("this class pulls in
more/heavier library machinery than that one"), not an absolute byte
count. For real numbers, measure the actual build's map file / `nm`
`--size-sort` / `size` output and overwrite the reference table's KB
column with measured values — the Excel output is editable for exactly
this reason.

Usage:
    python3 tools/library_feature_rom_audit.py

    Copy the files you want analyzed (any format, any subfolder
    structure) into library_feature_rom_audit_target/, then run the
    script. It writes library_feature_rom_audit_output/result.csv and
    result.xlsx (needs openpyxl for the latter) and prints a summary
    to stdout.

    --target / --output / --csv / --xlsx let you override the default
    folders for one-off runs; see `--help`.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from coverage_audit import (  # noqa: E402  (reuse the tested mini-parser)
    _CLASS_RE,
    find_matching_brace,
    strip_comments_and_literals,
    strip_preprocessor_directives,
)
from _io_layout import iter_all_files, output_dir_for, target_dir_for  # noqa: E402

# ---------------------------------------------------------------------------
# Feature reference table: name -> (regex, impact tier, rough KB, note)
#
# All KB figures are *order-of-magnitude guidelines*, not measurements.
# See the module docstring.
# ---------------------------------------------------------------------------

Feature = tuple  # (pattern: re.Pattern, tier: str, kb_low: float, kb_high: float, note: str)

_WB = r"(?<![\w:])"   # left word/scope boundary
_WB_END = r"(?![\w])"


def _tok(s: str) -> re.Pattern:
    return re.compile(_WB + re.escape(s) + _WB_END)


FEATURES: dict[str, tuple[re.Pattern, str, float, float, str]] = {
    "iostream(cout/cin/cerr)": (
        re.compile(_WB + r"std::(cout|cin|cerr|clog)" + _WB_END), "VeryLarge", 10, 30,
        "iostreamの静的初期化・locale関連コードが芋づる式にリンクされる。組み込みでは避けるのが定石。",
    ),
    "stringstream/fstream": (
        re.compile(_WB + r"std::(stringstream|ostringstream|istringstream|fstream|ifstream|ofstream)" + _WB_END),
        "Large", 8, 20, "iostream系と同様の静的初期化コストに加え、バッファ管理コードが乗る。",
    ),
    "例外(try/catch/throw)": (
        re.compile(r"\b(try|catch|throw)\b"), "Large", 5, 20,
        "unwindテーブル等がプログラム全体に波及する。throw箇所が増えるほど増加。",
    ),
    "RTTI(typeid/dynamic_cast)": (
        re.compile(r"\b(typeid\s*\(|dynamic_cast\s*<)"), "Medium", 0.5, 2,
        "クラスごとにtype_info生成・比較コードが追加される。",
    ),
    "std::regex": (
        _tok("std::regex"), "VeryLarge", 50, 100,
        "状態機械テンプレートが極めて大きい。組み込みでは基本的に非推奨。",
    ),
    "std::string": (
        _tok("std::string") ,"Medium", 2, 5,
        "SSO・アロケータまわりのコードが初回インスタンス化される。",
    ),
    "std::vector": (
        re.compile(_WB + r"std::vector\s*<"), "Medium", 1, 3,
        "要素型ごとにテンプレートが実体化される（型数に比例して増加）。",
    ),
    "std::map/std::set": (
        re.compile(_WB + r"std::(map|set|multimap|multiset)\s*<"), "Large", 3, 6,
        "赤黒木のコードが型ごとに実体化される。",
    ),
    "std::unordered_map/set": (
        re.compile(_WB + r"std::unordered_(map|set|multimap|multiset)\s*<"), "Large", 4, 8,
        "ハッシュテーブル実装は木構造より大きくなりがち。",
    ),
    "std::shared_ptr/weak_ptr": (
        re.compile(_WB + r"std::(shared_ptr|weak_ptr|make_shared)\s*[<(]"), "Medium", 1, 2,
        "コントロールブロック＋atomic参照カウントのコードが乗る。",
    ),
    "std::unique_ptr": (
        re.compile(_WB + r"std::(unique_ptr|make_unique)\s*[<(]"), "Small", 0.2, 0.5,
        "ほぼゼロオーバーヘッドだが、デリータ呼び出し分のコードは乗る。",
    ),
    "std::function": (
        re.compile(_WB + r"std::function\s*<"), "Medium", 1, 2,
        "型消去のためのvtable相当コードが実体化される。",
    ),
    "std::thread": (
        _tok("std::thread"), "VeryLarge", 10, 20,
        "スレッドランタイムのリンクが必要になる。すでにリンク済みなら限界コストは下がる。",
    ),
    "std::mutex/condition_variable": (
        re.compile(_WB + r"std::(mutex|recursive_mutex|condition_variable|lock_guard|unique_lock)\s*[<(]?"),
        "Small", 0.2, 1,
        "多くの場合OSプリミティブ（VxWorxセマフォ等）への薄いラッパー。",
    ),
    "std::chrono": (
        _tok("std::chrono") if False else re.compile(r"std::chrono::"), "Small", 0.1, 0.5,
        "ほぼコンパイル時計算で、実行時コストは小さい。",
    ),
    "template(汎用)": (
        re.compile(r"\btemplate\s*<"), "Variable", 0, 0,
        "実体化される型の組み合わせ数に比例してコード量が増える（コンパイル時に決まる）。",
    ),
    "virtual関数": (
        re.compile(r"\bvirtual\b"), "Small", 0.05, 0.05,
        "vtableエントリ1つあたりの目安。クラスごとのvtable自体の固定コストは別途乗る。",
    ),
    "動的確保(new/delete)": (
        re.compile(r"\b(new|delete)\b"), "Small", 0.5, 1,
        "初回はヒープアロケータ本体のリンクコストが乗る（以降は限界コストが下がる）。",
    ),
    "std::variant": (
        re.compile(_WB + r"std::variant\s*<"), "Medium", 1, 3,
        "visitディスパッチのコードが型の組み合わせごとに実体化される。",
    ),
    "std::optional": (
        re.compile(_WB + r"std::optional\s*<"), "Small", 0.2, 0.5,
        "軽量。ムーブ/コピーコンストラクタ相当が乗る程度。",
    ),
    "<algorithm>系(sort/find等)": (
        re.compile(_WB + r"std::(sort|find|find_if|for_each|transform|accumulate|copy|remove_if|count_if)\s*\("),
        "Small", 0.5, 2,
        "呼び出す型・イテレータの組み合わせごとに実体化される。",
    ),
    "文字列変換(to_string等)": (
        re.compile(_WB + r"std::(to_string|stoi|stol|stod|stof)\s*\("), "Small", 0.5, 1,
        "内部でlocale/フォーマットコードを引き込む場合がある。",
    ),
}

TIER_ORDER = {"Small": 1, "Medium": 2, "Large": 3, "VeryLarge": 4, "Variable": 0}

# ---------------------------------------------------------------------------
# Layer classification (same convention as coverage_audit.py)
# ---------------------------------------------------------------------------

LAYER_RULES = [
    ("core/storage/accessor", "Accessor"),
    ("core/storage/snapshot", "Snapshot"),
    ("core/storage", "Storage"),
    ("core/adapter", "Adapter"),
    ("core/capability", "Capability"),
    ("core/publisher", "Publisher"),
    ("core/store", "Store"),
    ("core/common", "Common"),
    ("include/rim_api.h", "C-API"),
    ("printer_a.h", "C-API"),
    ("products/printer_a", "Product(printer_a)"),
    ("products/common", "Product(common)"),
]


def classify_layer(path: Path) -> str:
    rel = path.as_posix()
    for fragment, label in LAYER_RULES:
        if fragment in rel:
            return label
    return "Other"


# ---------------------------------------------------------------------------
# Scanning
# ---------------------------------------------------------------------------

_INCLUDE_RE = re.compile(r'^\s*#\s*include\s*[<"]([^">]+)[">]', re.MULTILINE)


@dataclass
class ClassUsage:
    layer: str
    class_name: str
    source_file: str
    includes: list[str] = field(default_factory=list)
    feature_counts: dict[str, int] = field(default_factory=dict)

    @property
    def rom_estimate_kb(self) -> tuple[float, float]:
        low = high = 0.0
        for name, count in self.feature_counts.items():
            if count <= 0:
                continue
            _, tier, kb_low, kb_high, _ = FEATURES[name]
            if tier == "Variable":
                continue
            low += kb_low
            high += kb_high
        return (low, high)


def find_class_blocks(text: str) -> list[tuple[str, str, str]]:
    """Return [(kind, name, body_text)] for every top-level class/struct."""
    blocks = []
    for m in _CLASS_RE.finditer(text):
        kind = m.group(2)
        name = m.group(3)
        open_brace = m.end() - 1
        close_brace = find_matching_brace(text, open_brace)
        blocks.append((kind, name, text[open_brace + 1:close_brace]))
    return blocks


def scan_features(text: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for name, (pattern, *_rest) in FEATURES.items():
        n = len(pattern.findall(text))
        if n:
            counts[name] = n
    return counts


def scan_file(path: Path, target_dir: Path) -> list[ClassUsage]:
    raw = path.read_text(encoding="utf-8", errors="replace")
    includes = sorted(set(_INCLUDE_RE.findall(raw)))
    rel_path = path.relative_to(target_dir)
    layer = classify_layer(rel_path)
    rel = str(rel_path)

    text = strip_preprocessor_directives(strip_comments_and_literals(raw))

    results: list[ClassUsage] = []
    class_spans: list[tuple[int, int]] = []

    for m in _CLASS_RE.finditer(text):
        name = m.group(3)
        open_brace = m.end() - 1
        close_brace = find_matching_brace(text, open_brace)
        body = text[open_brace + 1:close_brace]
        class_spans.append((m.start(), close_brace))
        counts = scan_features(body)
        if counts:
            results.append(ClassUsage(
                layer=layer, class_name=name, source_file=rel,
                includes=includes, feature_counts=counts,
            ))

    # Anything at file scope, outside any class body, is attributed to a
    # synthetic "(file-level)" pseudo-class so free-function usage isn't
    # silently dropped.
    outside_chunks = []
    cursor = 0
    for start, end in sorted(class_spans):
        if start > cursor:
            outside_chunks.append(text[cursor:start])
        cursor = max(cursor, end + 1)
    outside_chunks.append(text[cursor:])
    outside_text = "\n".join(outside_chunks)
    outside_counts = scan_features(outside_text)
    if outside_counts:
        results.append(ClassUsage(
            layer=layer, class_name="(file-level)", source_file=rel,
            includes=includes, feature_counts=outside_counts,
        ))

    return results


def collect_from_target(target_dir: Path) -> list[ClassUsage]:
    usages: list[ClassUsage] = []
    for p in iter_all_files(target_dir):
        usages.extend(scan_file(p, target_dir))
    return usages


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

FEATURE_NAMES = list(FEATURES.keys())


def write_csv(usages: list[ClassUsage], out_path: Path) -> None:
    import csv
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Layer", "Class", "SourceFile", "ROM目安低(KB)", "ROM目安高(KB)"] + FEATURE_NAMES)
        for u in usages:
            low, high = u.rom_estimate_kb
            w.writerow([u.layer, u.class_name, u.source_file, f"{low:.2f}", f"{high:.2f}"]
                       + [u.feature_counts.get(name, "") for name in FEATURE_NAMES])


def write_xlsx(usages: list[ClassUsage], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FONT_NAME = "Arial"
    HEADER_FILL = PatternFill("solid", fgColor="203864")
    HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    TIER_FILL = {
        "Small": PatternFill("solid", fgColor="C6E0B4"),
        "Medium": PatternFill("solid", fgColor="FFE699"),
        "Large": PatternFill("solid", fgColor="F8CBAD"),
        "VeryLarge": PatternFill("solid", fgColor="FF8080"),
        "Variable": PatternFill("solid", fgColor="D9D9D9"),
    }

    wb = Workbook()

    # --- Sheet 1: reference table -----------------------------------
    ws1 = wb.active
    ws1.title = "凡例・機能一覧"
    ws1.sheet_view.showGridLines = False

    title_font = Font(name=FONT_NAME, size=16, bold=True, color="203864")
    ws1["B2"] = "C++標準ライブラリ機能 使用状況・ROM目安調査"
    ws1["B2"].font = title_font
    ws1["B3"] = ("注意: 以下のKB値は一般的な組み込みC++開発の経験則に基づく目安であり、実測値ではありません。"
                 "実際の値はコンパイラ・最適化オプション・リンカのdead code除去設定・ターゲットアーキテクチャに"
                 "強く依存します。正確な値が必要な場合は実ビルドのmapファイル/nm/sizeコマンドで実測し、"
                 "下表のKB列を実測値で上書きしてください（他シートの数式はこの表を参照して自動再計算されます）。")
    ws1["B3"].font = Font(name=FONT_NAME, size=9, italic=True, color="595959")
    ws1["B3"].alignment = Alignment(wrap_text=True)
    ws1.merge_cells("B3:H3")
    ws1.row_dimensions[3].height = 60

    headers = ["機能", "ROM影響度", "目安KB(下限)", "目安KB(上限)", "備考"]
    header_row = 5
    for i, h in enumerate(headers):
        c = ws1[f"{get_column_letter(2 + i)}{header_row}"]
        c.value = h
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER

    for i, name in enumerate(FEATURE_NAMES):
        pattern, tier, kb_low, kb_high, note = FEATURES[name]
        row = header_row + 1 + i
        values = [name, tier, kb_low, kb_high, note]
        for ci, v in enumerate(values):
            cell = ws1[f"{get_column_letter(2 + ci)}{row}"]
            cell.value = v
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == 4))
            if ci == 1:
                cell.fill = TIER_FILL.get(tier, PatternFill())

    ws1.column_dimensions["A"].width = 3
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 12
    ws1.column_dimensions["F"].width = 60
    feature_table_first_row = header_row + 1
    feature_table_last_row = header_row + len(FEATURE_NAMES)

    # --- Sheet 2: per-class matrix ------------------------------------
    ws2 = wb.create_sheet("クラス別使用状況")
    ws2.sheet_view.showGridLines = False

    fixed_headers = ["レイヤ", "クラス", "ソースファイル", "推定ROM影響 下限(KB)", "推定ROM影響 上限(KB)"]
    all_headers = fixed_headers + FEATURE_NAMES
    for i, h in enumerate(all_headers):
        c = ws2[f"{get_column_letter(i + 1)}1"]
        c.value = h
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws2.row_dimensions[1].height = 60

    n_fixed = len(fixed_headers)
    feature_col_start = n_fixed + 1  # 1-based

    # Rows 2/3: per-feature "low KB" / "high KB" pulled from Sheet1 via
    # INDEX/MATCH, used by the SUMPRODUCT formulas below. Small-font
    # helper rows, not meant to be read directly.
    ws2.cell(row=2, column=1, value="(目安KB下限 参照行)").font = Font(name=FONT_NAME, size=8, italic=True, color="808080")
    ws2.cell(row=3, column=1, value="(目安KB上限 参照行)").font = Font(name=FONT_NAME, size=8, italic=True, color="808080")
    for i, name in enumerate(FEATURE_NAMES):
        col = feature_col_start + i
        col_letter = get_column_letter(col)
        for helper_row, sheet1_col in ((2, "D"), (3, "E")):
            formula = (
                f"=INDEX('凡例・機能一覧'!${sheet1_col}${feature_table_first_row}:${sheet1_col}${feature_table_last_row},"
                f"MATCH({col_letter}$1,'凡例・機能一覧'!$B${feature_table_first_row}:$B${feature_table_last_row},0))"
            )
            cell = ws2[f"{col_letter}{helper_row}"]
            cell.value = formula
            cell.font = Font(name=FONT_NAME, size=8, italic=True, color="808080")

    widths = [16, 26, 40, 12, 12] + [14] * len(FEATURE_NAMES)
    for i, w in enumerate(widths):
        ws2.column_dimensions[get_column_letter(i + 1)].width = w

    data_first_row = 4
    for ridx, u in enumerate(usages):
        row = data_first_row + ridx
        feat_first_letter = get_column_letter(feature_col_start)
        feat_last_letter = get_column_letter(feature_col_start + len(FEATURE_NAMES) - 1)
        low_formula = (
            f"=SUMPRODUCT(--({feat_first_letter}{row}:{feat_last_letter}{row}>0),"
            f"{feat_first_letter}$2:{feat_last_letter}$2)"
        )
        high_formula = (
            f"=SUMPRODUCT(--({feat_first_letter}{row}:{feat_last_letter}{row}>0),"
            f"{feat_first_letter}$3:{feat_last_letter}$3)"
        )

        values = [u.layer, u.class_name, u.source_file, low_formula, high_formula]
        for ci, v in enumerate(values[:n_fixed]):
            cell = ws2[f"{get_column_letter(ci + 1)}{row}"]
            cell.value = v
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == 2))

        for i, name in enumerate(FEATURE_NAMES):
            col = feature_col_start + i
            cell = ws2[f"{get_column_letter(col)}{row}"]
            count = u.feature_counts.get(name, "")
            cell.value = count if count != "" else None
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=9)
            cell.alignment = Alignment(horizontal="center")
            if count:
                _, tier, *_ = FEATURES[name]
                cell.fill = TIER_FILL.get(tier, PatternFill())

    ws2.freeze_panes = "D4"
    last_row = data_first_row + len(usages) - 1
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{max(last_row, data_first_row)}"

    # --- Sheet 3: includes per file ------------------------------------
    ws3 = wb.create_sheet("ファイル別Include一覧")
    ws3.sheet_view.showGridLines = False
    ws3["A1"] = "ファイル"
    ws3["B1"] = "レイヤ"
    ws3["C1"] = "#include一覧(山括弧/引用符とも)"
    for col in "ABC":
        c = ws3[f"{col}1"]
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 90

    seen_files: dict[str, tuple[str, list[str]]] = {}
    for u in usages:
        seen_files[u.source_file] = (u.layer, u.includes)
    for ridx, (fname, (layer, includes)) in enumerate(sorted(seen_files.items())):
        row = 2 + ridx
        ws3[f"A{row}"] = fname
        ws3[f"B{row}"] = layer
        ws3[f"C{row}"] = ", ".join(includes)
        for col in "ABC":
            ws3[f"{col}{row}"].border = BORDER
            ws3[f"{col}{row}"].font = Font(name=FONT_NAME, size=9)
            ws3[f"{col}{row}"].alignment = Alignment(wrap_text=(col == "C"), vertical="top")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def print_summary(usages: list[ClassUsage]) -> None:
    from collections import Counter

    print(f"# 機能使用が検出されたクラス/ファイルレベル単位: {len(usages)}")
    feature_hits = Counter()
    for u in usages:
        for name in u.feature_counts:
            feature_hits[name] += 1
    print()
    print(f"{'機能':<28} {'検出クラス数':>10} {'ROM影響度':>10}")
    for name, n in feature_hits.most_common():
        tier = FEATURES[name][1]
        print(f"{name:<28} {n:>10} {tier:>10}")

    ranked = sorted(usages, key=lambda u: u.rom_estimate_kb[1], reverse=True)
    print()
    print("# 推定ROM影響(目安・上限KB)が大きいクラス Top10")
    for u in ranked[:10]:
        low, high = u.rom_estimate_kb
        print(f"  - [{u.layer}] {u.class_name} ({u.source_file}): 目安 {low:.1f}〜{high:.1f} KB")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--target", type=Path, default=None,
                         help="Folder to scan recursively (default: library_feature_rom_audit_target/ next to this script)")
    parser.add_argument("--output", type=Path, default=None,
                         help="Folder to write results into (default: library_feature_rom_audit_output/ next to this script)")
    parser.add_argument("--csv", type=Path, default=None, help="Override CSV output path (default: <output>/result.csv)")
    parser.add_argument("--xlsx", type=Path, default=None, help="Override Excel output path (default: <output>/result.xlsx, needs openpyxl)")
    args = parser.parse_args()

    target_dir = args.target.resolve() if args.target else target_dir_for(__file__)
    output_dir = args.output.resolve() if args.output else output_dir_for(__file__)

    if not target_dir.is_dir():
        print(f"Target folder not found: {target_dir}", file=sys.stderr)
        return 1

    print(f"# 解析対象フォルダ: {target_dir}")
    usages = collect_from_target(target_dir)
    usages.sort(key=lambda u: (u.layer, u.class_name))

    print_summary(usages)

    csv_path = args.csv if args.csv else output_dir / "result.csv"
    write_csv(usages, csv_path)
    print(f"\nCSV written: {csv_path}")

    xlsx_path = args.xlsx if args.xlsx else output_dir / "result.xlsx"
    try:
        write_xlsx(usages, xlsx_path)
        print(f"Excel written: {xlsx_path}")
    except ImportError:
        print("openpyxl is not installed; skipping Excel output "
              "(pip install openpyxl)", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
