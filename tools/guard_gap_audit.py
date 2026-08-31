#!/usr/bin/env python3
"""guard_gap_audit.py — mechanical survey of missing abnormal-case guards.

For every method body found under a source tree (inline class methods
in headers, out-of-line `Class::Method(...) { ... }` definitions in
.cpp files, and free functions), this script checks 13 categories of
defensive/abnormal-case handling this codebase is known to rely on
(null checks, uninitialized-state guards, not-found checks, type
mismatch checks, size/range checks, idempotency guards, ...), and
reports which methods look like they *should* have a given guard
(based on their signature/body shape) but apparently don't.

This is a regex/heuristic tool, NOT a data-flow or control-flow
analyzer. "Applicable" (a category's trigger condition matched) does
not prove the guard is actually required, and "guarded" only proves
*some* guard-shaped pattern exists somewhere in the method body — not
that it protects the specific risky operation. Treat every row as a
lead to manually verify, not a confirmed defect. Each category below
carries a rough confidence label (高/中/低) reflecting how directly
its heuristic maps to the real risk.

Usage:
    python3 tools/guard_gap_audit.py                      # auto-detect latest RevNNN_FW, summary to stdout
    python3 tools/guard_gap_audit.py --root Rev755_FW      # target a specific folder
    python3 tools/guard_gap_audit.py --xlsx out.xlsx       # detailed Excel
    python3 tools/guard_gap_audit.py --csv out.csv         # detailed CSV
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from coverage_audit import (  # noqa: E402  (reuse the tested mini-parser)
    _CLASS_RE,
    find_matching_brace,
    find_matching_paren,
    flatten_wrappers,
    isolate_access_labels,
    split_by_access,
    split_header_and_body,
    split_statements,
    strip_comments_and_literals,
    strip_preprocessor_directives,
)

REPO_ROOT = Path(__file__).resolve().parent.parent

DEFAULT_EXCLUDE_DIR_NAMES = {"test", ".vscode", ".devcontainer", ".github"}

# ---------------------------------------------------------------------------
# Rev folder auto-detection
# ---------------------------------------------------------------------------

_REV_DIR_RE = re.compile(r"^Rev(\d+)_FW$")


def find_latest_rev_dir(root: Path) -> Path | None:
    candidates = []
    for p in root.iterdir():
        if not p.is_dir():
            continue
        m = _REV_DIR_RE.match(p.name)
        if m:
            candidates.append((int(m.group(1)), p))
    if not candidates:
        return None
    candidates.sort(key=lambda t: t[0])
    return candidates[-1][1]


# ---------------------------------------------------------------------------
# Layer classification (best-effort, path-fragment based)
# ---------------------------------------------------------------------------

LAYER_RULES = [
    ("core/adapter", "Adapter"),
    ("core/capability", "Capability"),
    ("core/publisher", "Publisher"),
    ("core/storage", "Storage"),
    ("core/store", "Store"),
    ("core/common", "Common"),
    ("rim_api", "C-API"),
    ("products", "Product"),
]


def classify_layer(path: Path) -> str:
    rel = path.as_posix()
    for fragment, label in LAYER_RULES:
        if fragment in rel:
            return label
    return "Other"


# ---------------------------------------------------------------------------
# Small helpers on top of coverage_audit's primitives
# ---------------------------------------------------------------------------

_NAME_PAREN_RE = re.compile(r"([~]?\w+)\s*\(")
_QUALIFIED_DEF_RE = re.compile(r"(~?\w+)\s*::\s*(~?\w+)\s*\(")


def split_top_level(text: str, sep: str = ",") -> list[str]:
    """Split on a separator, ignoring occurrences nested inside (), <>, []."""
    parts = []
    depth_paren = depth_angle = depth_bracket = 0
    start = 0
    for i, c in enumerate(text):
        if c == "(":
            depth_paren += 1
        elif c == ")":
            depth_paren = max(0, depth_paren - 1)
        elif c == "<":
            depth_angle += 1
        elif c == ">":
            depth_angle = max(0, depth_angle - 1)
        elif c == "[":
            depth_bracket += 1
        elif c == "]":
            depth_bracket = max(0, depth_bracket - 1)
        elif c == sep and depth_paren == depth_angle == depth_bracket == 0:
            parts.append(text[start:i])
            start = i + 1
    tail = text[start:]
    if tail.strip() or parts:
        parts.append(tail)
    return [p for p in parts if p.strip()]


@dataclass
class MethodBody:
    layer: str
    class_name: str
    method_name: str
    params: str
    body: str
    source_file: str


def extract_inline_class_methods(text: str, source_file: str, layer: str) -> list[MethodBody]:
    results = []
    for m in _CLASS_RE.finditer(text):
        kind = m.group(2)
        cname = m.group(3)
        open_brace = m.end() - 1
        close_brace = find_matching_brace(text, open_brace)
        raw_body = text[open_brace + 1:close_brace]
        default_access = "public" if kind == "struct" else "private"
        body = isolate_access_labels(raw_body)
        for _access, stmt in split_by_access(split_statements(body), default_access):
            header, has_body = split_header_and_body(stmt)
            if not has_body:
                continue
            header = header.strip()
            if not header or "(" not in header:
                continue
            matches = list(_NAME_PAREN_RE.finditer(header))
            if not matches:
                continue
            name_m = matches[0]
            mname = name_m.group(1)
            open_paren = header.index("(", name_m.start())
            close_paren = find_matching_paren(header, open_paren)
            params = header[open_paren + 1:close_paren]
            # body of this statement = everything after the header's own
            # opening brace, up to (but excluding) the statement's closing brace
            stripped = stmt.rstrip()
            body_start = len(header)
            inner = stripped[body_start:]
            first_brace = inner.find("{")
            if first_brace == -1:
                continue
            method_body_text = inner[first_brace + 1:-1] if inner.endswith("}") else inner[first_brace + 1:]
            results.append(MethodBody(layer, cname, mname, params, method_body_text, source_file))
    return results


def extract_out_of_class_methods(text: str, source_file: str, layer: str) -> list[MethodBody]:
    results = []
    flat = flatten_wrappers(text)
    for stmt in split_statements(flat):
        header, has_body = split_header_and_body(stmt)
        if not has_body:
            continue
        header = header.strip()
        m = _QUALIFIED_DEF_RE.search(header)
        if not m:
            continue
        cname, mname = m.group(1), m.group(2)
        open_paren = header.index("(", m.end() - 1)
        close_paren = find_matching_paren(header, open_paren)
        params = header[open_paren + 1:close_paren]
        stripped = stmt.rstrip()
        inner = stripped[len(header):]
        first_brace = inner.find("{")
        if first_brace == -1:
            continue
        method_body_text = inner[first_brace + 1:-1] if inner.endswith("}") else inner[first_brace + 1:]
        results.append(MethodBody(layer, cname, mname, params, method_body_text, source_file))
    return results


def extract_free_functions(text: str, source_file: str, layer: str) -> list[MethodBody]:
    """Top-level functions with a body that are *not* `Class::Method` and
    not inside a class (already handled by the two extractors above)."""
    results = []
    flat = flatten_wrappers(text)
    for stmt in split_statements(flat):
        header, has_body = split_header_and_body(stmt)
        if not has_body:
            continue
        header = header.strip()
        if not header or "::" in header:
            continue  # handled by extract_out_of_class_methods
        if _CLASS_RE.search(stmt):
            continue  # this statement IS a class definition, not a function
        matches = list(_NAME_PAREN_RE.finditer(header))
        if not matches:
            continue
        mname = matches[0].group(1)
        if mname.lower() in ("if", "for", "while", "switch", "namespace", "struct", "class"):
            continue
        open_paren = header.index("(", matches[0].start())
        close_paren = find_matching_paren(header, open_paren)
        params = header[open_paren + 1:close_paren]
        stripped = stmt.rstrip()
        inner = stripped[len(header):]
        first_brace = inner.find("{")
        if first_brace == -1:
            continue
        method_body_text = inner[first_brace + 1:-1] if inner.endswith("}") else inner[first_brace + 1:]
        results.append(MethodBody(layer, f"(free fn: {Path(source_file).name})", mname, params, method_body_text, source_file))
    return results


def scan_file(path: Path, repo_root: Path) -> list[MethodBody]:
    raw = path.read_text(encoding="utf-8", errors="replace")
    layer = classify_layer(path)
    rel = str(path.relative_to(repo_root))
    text = strip_preprocessor_directives(strip_comments_and_literals(raw))

    methods = extract_inline_class_methods(text, rel, layer)
    methods += extract_out_of_class_methods(text, rel, layer)
    methods += extract_free_functions(text, rel, layer)
    return methods


def collect(source_dirs: list[Path], repo_root: Path) -> list[MethodBody]:
    files: list[Path] = []
    for d in source_dirs:
        if not d.is_dir():
            continue
        for pattern in ("*.hpp", "*.h", "*.cpp"):
            for p in d.rglob(pattern):
                if any(part in DEFAULT_EXCLUDE_DIR_NAMES for part in p.parts):
                    continue
                files.append(p)

    methods: list[MethodBody] = []
    for p in sorted(set(files)):
        methods.extend(scan_file(p, repo_root))
    return methods


# ---------------------------------------------------------------------------
# Guard categories
# ---------------------------------------------------------------------------

NULLPTR_CHECK_RE = re.compile(r"(==\s*nullptr|!=\s*nullptr|nullptr\s*==)")
BARE_NEGATION_IF_RE = re.compile(r"if\s*\(\s*!\s*\w+\s*\)")
NEGATION_IF_CALL_RE = re.compile(r"if\s*\(\s*!")


def guard_nullptr_present(body: str) -> bool:
    return bool(NULLPTR_CHECK_RE.search(body)) or bool(BARE_NEGATION_IF_RE.search(body))


def has_pointer_param(params: str) -> bool:
    for p in split_top_level(params):
        type_part = p.split("=")[0]
        if "*" in type_part:
            return True
    return False


def has_size_param(params: str) -> bool:
    for p in split_top_level(params):
        if re.search(r"\bsize_t\b", p):
            return True
        if re.search(r"(?i)\b(count|size|len|length|num)\b", p):
            return True
    return False


@dataclass
class GuardCategory:
    key: str
    label: str
    confidence: str
    description: str
    trigger: "callable"
    guard: "callable"


def _trigger_new(params, body, mname, cname):
    return bool(re.search(r"\bnew\s+[A-Za-z_]", body)) and "make_unique" not in body and "make_shared" not in body


def _guard_new(params, body, mname, cname):
    return bool(re.search(r"\bdelete\b|unique_ptr|make_unique|\.release\s*\(\)|shared_ptr", body))


def _trigger_discard(params, body, mname, cname):
    return bool(re.search(r"\.(Push|TryPop|WaitAndPop|Dispatch)\s*\(", body))


def _guard_discard(params, body, mname, cname):
    bare_re = re.compile(
        r"^[ \t]*[\w:.>\[\]-]+\.(Push|TryPop|WaitAndPop|Dispatch)\s*\([^;]*\)\s*;\s*$",
        re.MULTILINE,
    )
    # "guarded" == no bare/discarded-result call site found
    return not bool(bare_re.search(body))


GUARD_CATEGORIES: list[GuardCategory] = [
    GuardCategory(
        "null_param", "NULLポインタチェック", "高",
        "ポインタ型引数を受け取るメソッドで、nullptrチェックが本体のどこかにあるか。",
        lambda p, b, m, c: has_pointer_param(p),
        lambda p, b, m, c: guard_nullptr_present(b),
    ),
    GuardCategory(
        "g_context", "未初期化状態チェック(g_context)", "高",
        "g_contextを参照するメソッドで、!g_context等のガードがあるか。",
        lambda p, b, m, c: "g_context" in b,
        # NOTE: must not match "!g_context->foo(...)" -- that '!' negates the
        # whole call expression, not g_context itself, and is a real gap
        # this category exists to catch (confirmed against RIM_GetNotification).
        lambda p, b, m, c: bool(re.search(r"!\s*g_context\b(?!\s*(->|\.))|g_context\s*==\s*nullptr", b)),
    ),
    GuardCategory(
        "find_not_found", "ID未登録チェック(Find系)", "中",
        "Find*()呼び出しがあるメソッドで、nullptrチェックが本体のどこかにあるか"
        "(呼び出し直後かは見ておらず、body全体での相関のみ)。",
        lambda p, b, m, c: bool(re.search(r"\bFind\w*\s*\(", b)),
        lambda p, b, m, c: guard_nullptr_present(b),
    ),
    GuardCategory(
        "type_mismatch", "型不一致チェック(Get/TryGet系)", "中",
        "RIMValueAccessor::Get系/TryGet系の呼び出しがあるメソッドで、"
        "戻り値を否定形if(!...)でチェックしているか。",
        lambda p, b, m, c: bool(re.search(r"RIMValueAccessor::\w+\s*\(|\bTryGet\w*\s*\(", b)),
        lambda p, b, m, c: bool(NEGATION_IF_CALL_RE.search(b)),
    ),
    GuardCategory(
        "size_range", "サイズ・範囲外チェック", "低",
        "size/count/len系の引数を持つメソッドで、0や上限との比較があるか。",
        lambda p, b, m, c: has_size_param(p),
        lambda p, b, m, c: bool(re.search(r"==\s*0\b|>\s*k?\w*(Max|Count|Size)\b|>=\s*k?\w*(Max|Count|Size)\b", b)),
    ),
    GuardCategory(
        "duplicate_init", "重複登録・多重初期化チェック", "中",
        "Create/Initialize/Start/Register/Subscribe系メソッドで、"
        "本体冒頭付近にif-return形の早期リターンがあるか。",
        lambda p, b, m, c: bool(re.match(r"(?i)^(create|initialize|start|register|subscribe)", m)),
        lambda p, b, m, c: bool(re.search(r"if\s*\([^)]*\)\s*\{?\s*return", b[:250])),
    ),
    GuardCategory(
        "unset_data", "未設定データへの参照チェック", "低",
        "Get/TryGet/Find系メソッド(命名ベース)で、否定形ifやnullptrチェックがあるか"
        "(型不一致チェック/ID未登録チェックと一部重複する可能性あり)。",
        lambda p, b, m, c: bool(re.match(r"(?i)^(get|tryget|find)", m)),
        lambda p, b, m, c: guard_nullptr_present(b) or bool(NEGATION_IF_CALL_RE.search(b)),
    ),
    GuardCategory(
        "queue_shutdown", "キューShutdown後操作チェック", "中",
        "Push/Pop系メソッドで、shutdown_フラグ/Shutdown()を参照しているか。",
        lambda p, b, m, c: bool(re.match(r"(?i)^(push|pop|trypop|waitandpop|enqueue|dequeue)", m)),
        lambda p, b, m, c: bool(re.search(r"(?i)shutdown_|shutdown\s*\(", b)),
    ),
    GuardCategory(
        "double_start_stop", "二重start/stopガード", "高",
        "start/stop/runという名前のメソッドで、running_系フラグのifガードがあるか。",
        lambda p, b, m, c: m.lower() in ("start", "stop", "run"),
        lambda p, b, m, c: bool(re.search(r"(?i)if\s*\([^)]*running", b)),
    ),
    GuardCategory(
        "ownership", "リソース所有権(二重解放/未解放)", "低",
        "生newを使っているメソッドで、delete/unique_ptr等の対応する解放コードがあるか。",
        _trigger_new,
        _guard_new,
    ),
    GuardCategory(
        "callback_null", "コールバック関数ポインタのnullチェック", "中",
        "store/diff/normalize等のコールバック風メンバを呼んでいるメソッドで、"
        "nullptrチェックが本体のどこかにあるか。",
        lambda p, b, m, c: bool(re.search(r"->\s*(store|diff|normalize)\s*\(|\bcallback\s*\(", b)),
        lambda p, b, m, c: guard_nullptr_present(b),
    ),
    GuardCategory(
        "discarded_return", "戻り値の握りつぶし", "中",
        "Push/TryPop/WaitAndPop/Dispatchの戻り値を、if/代入で使わず捨てている箇所がないか"
        "(1箇所でも捨てていれば「ガードなし」判定)。",
        _trigger_discard,
        _guard_discard,
    ),
    GuardCategory(
        "enum_cast", "不正なenum値・範囲外キャスト", "低",
        "Type/Method/Trigger/Status系のenumへstatic_castしているメソッドで、"
        "switchのdefault:またはassert()があるか。",
        lambda p, b, m, c: bool(re.search(r"static_cast\s*<\s*(rim::)?\w*(Type|Method|Trigger|Status)\w*\s*>", b)),
        lambda p, b, m, c: bool(re.search(r"\bdefault\s*:|assert\s*\(", b)),
    ),
]

CATEGORY_BY_KEY = {c.key: c for c in GUARD_CATEGORIES}


@dataclass
class MethodResult:
    layer: str
    class_name: str
    method_name: str
    source_file: str
    status: dict[str, str] = field(default_factory=dict)  # key -> "GUARDED"/"UNGUARDED"/"N/A"


def analyze(methods: list[MethodBody]) -> list[MethodResult]:
    results = []
    for mb in methods:
        r = MethodResult(mb.layer, mb.class_name, mb.method_name, mb.source_file)
        any_relevant = False
        for cat in GUARD_CATEGORIES:
            if cat.trigger(mb.params, mb.body, mb.method_name, mb.class_name):
                any_relevant = True
                guarded = cat.guard(mb.params, mb.body, mb.method_name, mb.class_name)
                r.status[cat.key] = "GUARDED" if guarded else "UNGUARDED"
            else:
                r.status[cat.key] = "N/A"
        if any_relevant:
            results.append(r)
    return results


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def write_csv(results: list[MethodResult], out_path: Path) -> None:
    import csv
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Layer", "Class", "Method", "SourceFile", "未ガード件数"] + [c.label for c in GUARD_CATEGORIES])
        for r in results:
            n_ungrd = sum(1 for v in r.status.values() if v == "UNGUARDED")
            w.writerow([r.layer, r.class_name, r.method_name, r.source_file, n_ungrd]
                       + [r.status[c.key] for c in GUARD_CATEGORIES])


def write_xlsx(results: list[MethodResult], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FONT_NAME = "Arial"
    HEADER_FILL = PatternFill("solid", fgColor="203864")
    HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    STATUS_FILL = {
        "GUARDED": PatternFill("solid", fgColor="C6E0B4"),
        "UNGUARDED": PatternFill("solid", fgColor="F8CBAD"),
        "N/A": PatternFill("solid", fgColor="D9D9D9"),
    }
    STATUS_SYMBOL = {"GUARDED": "⭕", "UNGUARDED": "❌", "N/A": "ー"}
    CONF_FILL = {
        "高": PatternFill("solid", fgColor="C6E0B4"),
        "中": PatternFill("solid", fgColor="FFE699"),
        "低": PatternFill("solid", fgColor="F8CBAD"),
    }

    wb = Workbook()

    # --- Sheet 1: legend ---------------------------------------------
    ws1 = wb.active
    ws1.title = "凡例"
    ws1.sheet_view.showGridLines = False
    ws1["B2"] = "異常系ガード欠落調査"
    ws1["B2"].font = Font(name=FONT_NAME, size=16, bold=True, color="203864")
    ws1["B3"] = ("注意: これは正規表現ベースの機械的ヒューリスティックであり、データフロー解析ではありません。"
                 "「該当あり」はメソッドの見た目(引数の型・呼んでいる関数名等)から見て、そのガードが"
                 "必要そうだと判定しただけで、実際に必要か・十分かは個別に確認してください。"
                 "「ガードなし」はメソッド本体のどこにもそれらしいガード表現が見つからなかったことを示します"
                 "(該当する処理の直後に限定したチェックではありません)。")
    ws1["B3"].font = Font(name=FONT_NAME, size=9, italic=True, color="595959")
    ws1["B3"].alignment = Alignment(wrap_text=True)
    ws1.merge_cells("B3:G3")
    ws1.row_dimensions[3].height = 75

    headers = ["カテゴリ", "検出信頼度", "説明"]
    header_row = 5
    for i, h in enumerate(headers):
        c = ws1[f"{get_column_letter(2 + i)}{header_row}"]
        c.value = h
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    for i, cat in enumerate(GUARD_CATEGORIES):
        row = header_row + 1 + i
        vals = [cat.label, cat.confidence, cat.description]
        for ci, v in enumerate(vals):
            cell = ws1[f"{get_column_letter(2 + ci)}{row}"]
            cell.value = v
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == 2))
            if ci == 1:
                cell.fill = CONF_FILL.get(v, PatternFill())
    ws1.column_dimensions["A"].width = 3
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 75

    # --- Sheet 2: per-method matrix -----------------------------------
    ws2 = wb.create_sheet("メソッド別ガード状況")
    ws2.sheet_view.showGridLines = False
    fixed_headers = ["レイヤ", "クラス", "メソッド", "ソースファイル", "未ガード件数"]
    all_headers = fixed_headers + [c.label for c in GUARD_CATEGORIES]
    for i, h in enumerate(all_headers):
        c = ws2[f"{get_column_letter(i + 1)}1"]
        c.value = h
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws2.row_dimensions[1].height = 55

    n_fixed = len(fixed_headers)
    cat_col_start = n_fixed + 1
    widths = [14, 26, 30, 45, 12] + [16] * len(GUARD_CATEGORIES)
    for i, w in enumerate(widths):
        ws2.column_dimensions[get_column_letter(i + 1)].width = w

    for ridx, r in enumerate(results):
        row = 2 + ridx
        cat_first = get_column_letter(cat_col_start)
        cat_last = get_column_letter(cat_col_start + len(GUARD_CATEGORIES) - 1)
        ungrd_formula = f'=COUNTIF({cat_first}{row}:{cat_last}{row},"❌")'

        fixed_vals = [r.layer, r.class_name, r.method_name, r.source_file, ungrd_formula]
        for ci, v in enumerate(fixed_vals):
            cell = ws2[f"{get_column_letter(ci + 1)}{row}"]
            cell.value = v
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == 3))

        for i, cat in enumerate(GUARD_CATEGORIES):
            col = cat_col_start + i
            status = r.status[cat.key]
            cell = ws2[f"{get_column_letter(col)}{row}"]
            cell.value = STATUS_SYMBOL[status]
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=10, bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = STATUS_FILL[status]

    ws2.freeze_panes = "D2"
    last_row = max(2, len(results) + 1)
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{last_row}"

    # --- Sheet 3: summary ------------------------------------------------
    ws3 = wb.create_sheet("サマリー")
    ws3.sheet_view.showGridLines = False
    ws3["B2"] = "カテゴリ別サマリー"
    ws3["B2"].font = Font(name=FONT_NAME, size=14, bold=True, color="203864")
    headers3 = ["カテゴリ", "該当メソッド数", "ガードあり", "ガードなし", "ガードなし率"]
    hrow = 4
    for i, h in enumerate(headers3):
        c = ws3[f"{get_column_letter(2 + i)}{hrow}"]
        c.value = h
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    data_last_row = max(2, len(results) + 1)
    for i, cat in enumerate(GUARD_CATEGORIES):
        row = hrow + 1 + i
        col_letter = get_column_letter(cat_col_start + i)
        applicable_f = (f'=COUNTIF(\'メソッド別ガード状況\'!{col_letter}2:{col_letter}{data_last_row},"⭕")'
                        f'+COUNTIF(\'メソッド別ガード状況\'!{col_letter}2:{col_letter}{data_last_row},"❌")')
        guarded_f = f'=COUNTIF(\'メソッド別ガード状況\'!{col_letter}2:{col_letter}{data_last_row},"⭕")'
        unguarded_f = f'=COUNTIF(\'メソッド別ガード状況\'!{col_letter}2:{col_letter}{data_last_row},"❌")'
        ratio_f = f'=IF(C{row}=0,0,E{row}/C{row})'

        ws3[f"B{row}"] = cat.label
        ws3[f"C{row}"] = applicable_f
        ws3[f"D{row}"] = guarded_f
        ws3[f"E{row}"] = unguarded_f
        ws3[f"F{row}"] = ratio_f
        ws3[f"F{row}"].number_format = "0.0%"
        for col in "BCDEF":
            cell = ws3[f"{col}{row}"]
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=9)
            cell.alignment = Alignment(horizontal="center" if col != "B" else "left")

    total_row = hrow + 1 + len(GUARD_CATEGORIES)
    ws3[f"B{total_row}"] = "合計(全メソッド, 重複あり)"
    ws3[f"B{total_row}"].font = Font(name=FONT_NAME, size=9, bold=True)
    for col, src in (("C", "C"), ("D", "D"), ("E", "E")):
        top = hrow + 1
        bot = total_row - 1
        ws3[f"{col}{total_row}"] = f"=SUM({src}{top}:{src}{bot})"
        ws3[f"{col}{total_row}"].font = Font(name=FONT_NAME, size=9, bold=True)
        ws3[f"{col}{total_row}"].border = BORDER

    ws3.column_dimensions["A"].width = 3
    ws3.column_dimensions["B"].width = 32
    for col in "CDEF":
        ws3.column_dimensions[col].width = 14

    wb.save(out_path)


def print_summary(results: list[MethodResult]) -> None:
    print(f"# 該当メソッド数(いずれかのカテゴリが適用されたメソッド): {len(results)}")
    print()
    print(f"{'カテゴリ':<32} {'該当':>6} {'ガードあり':>10} {'ガードなし':>10}")
    for cat in GUARD_CATEGORIES:
        applicable = [r for r in results if r.status[cat.key] != "N/A"]
        guarded = sum(1 for r in applicable if r.status[cat.key] == "GUARDED")
        unguarded = sum(1 for r in applicable if r.status[cat.key] == "UNGUARDED")
        print(f"{cat.label:<32} {len(applicable):>6} {guarded:>10} {unguarded:>10}")

    print()
    print("# ガードなし件数が多いメソッド Top15")
    ranked = sorted(
        results,
        key=lambda r: sum(1 for v in r.status.values() if v == "UNGUARDED"),
        reverse=True,
    )
    for r in ranked[:15]:
        n = sum(1 for v in r.status.values() if v == "UNGUARDED")
        if n == 0:
            break
        missing = [CATEGORY_BY_KEY[k].label for k, v in r.status.items() if v == "UNGUARDED"]
        print(f"  - [{r.layer}] {r.class_name}::{r.method_name} ({r.source_file}): {n}件")
        for label in missing:
            print(f"      - {label}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--root", type=str, default=None,
                         help="Target folder relative to repo root (e.g. Rev755_FW). "
                              "Default: auto-detect the highest-numbered RevNNN_FW at repo root.")
    parser.add_argument("--csv", type=Path, default=None, help="Write full results as CSV")
    parser.add_argument("--xlsx", type=Path, default=None, help="Write full results as a formatted Excel file (needs openpyxl)")
    parser.add_argument("--repo-root", type=Path, default=REPO_ROOT, help="Override repo root (default: parent of tools/)")
    args = parser.parse_args()

    repo_root = args.repo_root.resolve()

    if args.root:
        target = repo_root / args.root
    else:
        target = find_latest_rev_dir(repo_root)
        if target is None:
            print("No RevNNN_FW folder found at repo root and --root not given.", file=sys.stderr)
            return 1
        print(f"# 自動検出した対象フォルダ: {target.relative_to(repo_root)}")

    if not target.is_dir():
        print(f"Target folder not found: {target}", file=sys.stderr)
        return 1

    methods = collect([target], repo_root)
    results = analyze(methods)
    results.sort(key=lambda r: (r.layer, r.class_name, r.method_name))

    print_summary(results)

    if args.csv:
        write_csv(results, args.csv)
        print(f"\nCSV written: {args.csv}")

    if args.xlsx:
        try:
            write_xlsx(results, args.xlsx)
            print(f"Excel written: {args.xlsx}")
        except ImportError:
            print("openpyxl is not installed; skipping --xlsx output "
                  "(pip install openpyxl)", file=sys.stderr)
            return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
