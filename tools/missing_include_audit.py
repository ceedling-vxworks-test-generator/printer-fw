#!/usr/bin/env python3
"""missing_include_audit.py — detect symbols used without a visible
#include (the "IdentityDiff used in Capabilities.hpp but never
included" class of bug that Copilot has been flagging in this repo).

Many headers in this codebase define `inline`/`inline constexpr`
free functions, constexpr variables, classes, structs, enums and
type aliases at namespace scope, and other headers use them without
including the header that actually defines them. This "works" today
only because *some* translation unit happens to include the defining
header earlier via an unrelated path (e.g. Product.hpp happens to
include DataItem/DataItems.hpp, which pulls in CommonDiffFunctions.hpp,
before it includes CapabilityItem/Capabilities.hpp, which uses
IdentityDiff without including it itself). That's a landmine: include
the header standalone, reorder the includes, or drop the header that
was accidentally providing the symbol, and it stops compiling.

This script builds, from the files under missing_include_audit_target/:

  1. a symbol table of every top-level (namespace-scope, not inside a
     class body) function/variable/class/struct/enum/using declared or
     defined in each file;
  2. a local-`#include "..."` graph, resolved first relative to each
     including file's own directory, then (since a flat copy of the
     repo may not preserve every compiler include-path) by matching
     basenames against every file actually present in the target
     folder;
  3. for every file, the transitive closure of files reachable via
     that include graph (plus the file itself).

It then flags every (file, symbol) pair where the file's stripped text
references a symbol name that IS in the global symbol table, whose
defining file is NOT in the file's own transitive include closure, and
that the file does not itself define. That is exactly the "relies on
someone else including it first" pattern.

This is a regex/heuristic tool, not a preprocessor or a real include
graph (macros, conditional compilation, and system/library headers are
out of scope; only local quoted #include "..." are followed). Treat
every row as a lead to confirm by actually trying a standalone/
reordered compile, not a proven defect. In particular:

  * A flagged symbol used only inside a member-function body may
    still be "accidentally fine" today the same way it is now (via
    transitive includes from wherever this header itself gets
    included) — flagging it doesn't mean it currently fails to build,
    only that it does not stand on its own.
  * Short/very common identifiers are excluded (see MIN_NAME_LEN) to
    keep false positives down, at the cost of possibly missing some
    real short-named collisions.
  * If a symbol name is defined in more than one file, the row is
    flagged only when *none* of those defining files are in the
    closure; the "SuggestedInclude" column picks the shortest resulting
    relative #include path among the candidates as a starting point,
    not a guaranteed single right answer.

Usage:
    python3 tools/missing_include_audit.py

    Copy the files you want analyzed (any format, any subfolder
    structure -- preserving the original relative directory layout is
    important here since it drives both #include resolution and layer
    classification) into missing_include_audit_target/, then run the
    script. It writes missing_include_audit_output/result.csv and
    result.xlsx (needs openpyxl for the latter) and prints a summary
    to stdout.

    --target / --output / --csv / --xlsx let you override the default
    folders for one-off runs; see `--help`.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from coverage_audit import (  # noqa: E402  (reuse the tested mini-parser)
    _CLASS_RE,
    _CONTROL_KEYWORDS,
    _NAME_PAREN_RE,
    find_matching_brace,
    flatten_wrappers,
    split_header_and_body,
    split_statements,
    strip_comments_and_literals,
    strip_preprocessor_directives,
)
from _io_layout import iter_all_files, output_dir_for, target_dir_for  # noqa: E402

MIN_NAME_LEN = 3

# Directory-path fragment -> display layer name, matched against the
# file's path relative to the target folder (same convention as the
# other tools/*.py scripts).
LAYER_RULES = [
    ("core/storage/accessor", "Accessor"),
    ("core/storage/snapshot", "Snapshot"),
    ("core/storage", "Storage"),
    ("core/adapter", "Adapter"),
    ("core/capability", "Capability"),
    ("core/publisher", "Publisher"),
    ("core/store", "Store"),
    ("core/common", "Common"),
    ("products/common", "Product(common)"),
    ("rim_api", "C-API"),
    ("products", "Product"),
]


def classify_layer(rel_path: Path) -> str:
    rel = rel_path.as_posix()
    for fragment, label in LAYER_RULES:
        if fragment in rel:
            return label
    return "Other"


# ---------------------------------------------------------------------------
# Symbol extraction: every namespace-scope function/variable/class/
# struct/enum/using this file declares or defines.
# ---------------------------------------------------------------------------

_ENUM_RE = re.compile(r"\benum\s+(?:class\s+|struct\s+)?(\w+)\s*(?::\s*[\w:]+\s*)?\{")
_FWD_DECL_RE = re.compile(r"\b(?:class|struct)\s+(\w+)\s*;")
_USING_ALIAS_RE = re.compile(r"^using\s+(\w+)\s*=")
_TRAILING_IDENT_RE = re.compile(r"([A-Za-z_]\w*)\s*(?:\[[^\]]*\])?\s*$")

# Basic type keywords / control-flow-ish tokens that can legitimately sit
# directly before a top-level "(" (e.g. a `typedef void (*Cb)(...)`
# fragment, or a bare macro invocation like `TEST(Suite, Case) { ... }`)
# without being a real declared symbol name.
_NON_SYMBOL_NAMES = {
    "void", "int", "char", "bool", "float", "double", "long", "short",
    "unsigned", "signed", "auto", "wchar_t", "size_t",
}

# File types that are known to produce false matches against the C++-ish
# regexes above (CMake's `set(...)`/`add_library(...)`, Markdown prose,
# etc.) without ever containing a real #include graph. Excluded from
# both the symbol table and the usage scan; everything else is still
# scanned regardless of extension, per the format-agnostic convention
# tools/_io_layout.py documents.
_NON_CPP_EXCLUDE_NAMES = {"cmakelists.txt"}
_NON_CPP_EXCLUDE_EXTS = {".md", ".txt", ".cmake", ".json", ".yml", ".yaml", ".xml", ".py", ".sh", ".rst"}


def _looks_like_cpp(path: Path) -> bool:
    if path.name.lower() in _NON_CPP_EXCLUDE_NAMES:
        return False
    if path.suffix.lower() in _NON_CPP_EXCLUDE_EXTS:
        return False
    return True


@dataclass
class Symbol:
    name: str
    kind: str  # Function / Variable / Class / Struct / Enum / Using
    file: Path


def _trailing_identifier(header: str) -> str | None:
    h = header.strip()
    if h.endswith("="):
        h = h[:-1].strip()
    m = _TRAILING_IDENT_RE.search(h)
    if not m:
        return None
    name = m.group(1)
    if name.lower() in _CONTROL_KEYWORDS or len(name) < MIN_NAME_LEN:
        return None
    return name


def collect_symbols(text: str, path: Path) -> list[Symbol]:
    symbols: list[Symbol] = []

    for m in _CLASS_RE.finditer(text):
        name = m.group(3)
        if len(name) >= MIN_NAME_LEN:
            kind = "Class" if m.group(2) == "class" else "Struct"
            symbols.append(Symbol(name, kind, path))

    for m in _ENUM_RE.finditer(text):
        name = m.group(1)
        if len(name) >= MIN_NAME_LEN:
            symbols.append(Symbol(name, "Enum", path))

    # A forward declaration (`class Foo;` with no body) is enough for a
    # file to use that type by pointer/reference without including its
    # real definition -- record it as a provider too, so a header that
    # forward-declares and then only uses a type that way isn't flagged
    # against itself.
    for m in _FWD_DECL_RE.finditer(text):
        name = m.group(1)
        if len(name) >= MIN_NAME_LEN:
            symbols.append(Symbol(name, "Class", path))

    flat = flatten_wrappers(text)
    for stmt in split_statements(flat):
        # A statement that opens with `class`/`struct { ... }` swallows
        # the whole member list as one blob -- skip it here, it (and
        # its top-level name) is already handled above; this also
        # keeps member declarations from being misread as free symbols.
        if _CLASS_RE.search(stmt) or _ENUM_RE.search(stmt):
            continue

        header, _has_body = split_header_and_body(stmt)
        header = header.strip()
        if not header:
            continue

        if re.match(r"^typedef\b", header):
            continue  # `typedef RET (*Name)(...)` etc. -- not reliably parseable here

        um = _USING_ALIAS_RE.match(header)
        if um:
            if len(um.group(1)) >= MIN_NAME_LEN:
                symbols.append(Symbol(um.group(1), "Using", path))
            continue

        matches = list(_NAME_PAREN_RE.finditer(header))
        if matches:
            prefix = header[:matches[0].start()]
            if prefix.rstrip().endswith("::"):
                continue  # out-of-line qualified member definition, e.g. Class::Method(...)
            if prefix.strip():
                # Something (a return type, `inline`, `static`, ...)
                # precedes the identifier -- a bare `NAME(...)` with
                # nothing before it is a macro/call-like statement
                # (gtest's `TEST(Suite, Case)`, CMake's `set(...)`),
                # not a declaration.
                name = matches[0].group(1)
                if name.lower() not in _CONTROL_KEYWORDS and name not in _NON_SYMBOL_NAMES and len(name) >= MIN_NAME_LEN:
                    symbols.append(Symbol(name, "Function", path))
            continue

        if re.search(r"\b(inline|constexpr|extern)\b", header):
            name = _trailing_identifier(header)
            if name and name not in _NON_SYMBOL_NAMES:
                symbols.append(Symbol(name, "Variable", path))

    return symbols


def collect_member_names(text: str) -> set[str]:
    """Names of methods/fields declared inside any class/struct body in
    this file. A member named e.g. `GetBool` sharing a name with some
    unrelated free function elsewhere is not a "use" of that free
    function -- it is a same-named member of this file's own class."""
    names: set[str] = set()
    for m in _CLASS_RE.finditer(text):
        open_brace = m.end() - 1
        close_brace = find_matching_brace(text, open_brace)
        for stmt in split_statements(text[open_brace + 1:close_brace]):
            header, _has_body = split_header_and_body(stmt)
            header = header.strip()
            if not header:
                continue
            matches = list(_NAME_PAREN_RE.finditer(header))
            if matches:
                names.add(matches[0].group(1))
    return names


# ---------------------------------------------------------------------------
# #include graph
# ---------------------------------------------------------------------------

_INCLUDE_QUOTED_RE = re.compile(r'^\s*#\s*include\s*"([^"]+)"', re.MULTILINE)


def resolve_local_includes(
    raw_text: str, including_file: Path, target_dir: Path, by_basename: dict[str, list[Path]],
) -> list[Path]:
    resolved: list[Path] = []
    for inc in _INCLUDE_QUOTED_RE.findall(raw_text):
        candidate = (including_file.parent / inc).resolve()
        if candidate.is_file():
            try:
                candidate.relative_to(target_dir)
                resolved.append(candidate)
                continue
            except ValueError:
                pass
        resolved.extend(by_basename.get(Path(inc).name, []))
    return resolved


def transitive_closure(start: Path, direct_includes: dict[Path, list[Path]]) -> set[Path]:
    seen = {start}
    stack = list(direct_includes.get(start, []))
    while stack:
        p = stack.pop()
        if p in seen:
            continue
        seen.add(p)
        stack.extend(direct_includes.get(p, []))
    return seen


# ---------------------------------------------------------------------------
# Finding assembly
# ---------------------------------------------------------------------------

_IDENT_TOKEN_RE = re.compile(r"\b[A-Za-z_]\w*\b")
_QUALIFIED_MEMBER_RE = re.compile(r"(?:::|\.|->)\s*([A-Za-z_]\w*)")


def used_identifiers(text: str) -> set[str]:
    """Every identifier token in text, except the right-hand side of a
    ::/./-> qualified access (Class::Method / obj.Method / obj->Method)
    -- a qualified member call must not be confused with an unrelated
    free function/variable that happens to share its short name."""
    masked = _QUALIFIED_MEMBER_RE.sub(
        lambda m: m.group(0)[:len(m.group(0)) - len(m.group(1))] + " " * len(m.group(1)), text,
    )
    return set(_IDENT_TOKEN_RE.findall(masked))


@dataclass
class Finding:
    layer: str
    file: str
    symbol: str
    kind: str
    defined_in: list[str]
    suggested_include: str
    note: str


def scan_target(target_dir: Path) -> list[Finding]:
    files = [p for p in iter_all_files(target_dir) if _looks_like_cpp(p)]

    raw_by_file: dict[Path, str] = {}
    stripped_by_file: dict[Path, str] = {}
    by_basename: dict[str, list[Path]] = {}
    for p in files:
        raw = p.read_text(encoding="utf-8", errors="replace")
        raw_by_file[p] = raw
        stripped_by_file[p] = strip_preprocessor_directives(strip_comments_and_literals(raw))
        by_basename.setdefault(p.name, []).append(p)

    symbol_table: dict[str, list[Symbol]] = {}
    for p in files:
        for sym in collect_symbols(stripped_by_file[p], p):
            symbol_table.setdefault(sym.name, []).append(sym)

    direct_includes: dict[Path, list[Path]] = {
        p: resolve_local_includes(raw_by_file[p], p, target_dir, by_basename) for p in files
    }
    closures: dict[Path, set[Path]] = {p: transitive_closure(p, direct_includes) for p in files}

    findings: list[Finding] = []
    for p in files:
        self_defined = {s.name for names in symbol_table.values() for s in names if s.file == p}
        member_names = collect_member_names(stripped_by_file[p])
        used = used_identifiers(stripped_by_file[p])
        closure = closures[p]
        rel_p = p.relative_to(target_dir)
        layer = classify_layer(rel_p)

        candidates = sorted(used & symbol_table.keys())
        for name in candidates:
            if name in self_defined or name in member_names:
                continue
            providers = symbol_table[name]
            provider_files = {s.file for s in providers}
            if provider_files & closure:
                continue

            best = min(provider_files, key=lambda f: len(os.path.relpath(f, start=p.parent)))
            suggested = Path(os.path.relpath(best, start=p.parent)).as_posix()
            note = "" if len(provider_files) == 1 else (
                f"複数箇所で定義されている可能性あり({len(provider_files)}件) — 最短経路を提案"
            )
            findings.append(Finding(
                layer=layer,
                file=str(rel_p),
                symbol=name,
                kind=providers[0].kind,
                defined_in=sorted(str(f.relative_to(target_dir)) for f in provider_files),
                suggested_include=f'#include "{suggested}"',
                note=note,
            ))

    return findings


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def write_csv(findings: list[Finding], out_path: Path) -> None:
    import csv
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Layer", "File", "Symbol", "Kind", "DefinedIn", "SuggestedInclude", "Note"])
        for r in findings:
            w.writerow([r.layer, r.file, r.symbol, r.kind, "; ".join(r.defined_in), r.suggested_include, r.note])


def write_xlsx(findings: list[Finding], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FONT_NAME = "Arial"
    HEADER_FILL = PatternFill("solid", fgColor="203864")
    HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    KIND_FILL = {
        "Function": PatternFill("solid", fgColor="FFE699"),
        "Variable": PatternFill("solid", fgColor="C6E0B4"),
        "Class": PatternFill("solid", fgColor="BDD7EE"),
        "Struct": PatternFill("solid", fgColor="BDD7EE"),
        "Enum": PatternFill("solid", fgColor="D9D9D9"),
        "Using": PatternFill("solid", fgColor="F8CBAD"),
    }

    wb = Workbook()

    # --- Sheet 1: legend --------------------------------------------------
    ws1 = wb.active
    ws1.title = "凡例"
    ws1.sheet_view.showGridLines = False
    ws1["B2"] = "未インクルードシンボル使用調査"
    ws1["B2"].font = Font(name=FONT_NAME, size=16, bold=True, color="203864")
    ws1["B3"] = (
        "注意: これは正規表現ベースのヒューリスティックであり、プリプロセッサや実際のコンパイラの"
        "include解決ではありません(マクロ・条件コンパイル・#include<...>形式のシステム/ライブラリ"
        "ヘッダは対象外。#include\"...\"のローカルインクルードのみを追跡します)。"
        "「該当あり」は、そのファイルが参照しているシンボル名が、ファイル自身の#includeを再帰的に"
        "辿った範囲(=推移的クロージャ)には見当たらず、別のファイルでしか定義/宣言されていない"
        "ことを示します。今はどこか別の翻訳単位が先にそのヘッダを読み込んでいるおかげでたまたま"
        "コンパイルが通っている可能性が高い箇所です。単体コンパイルやinclude順の入れ替えで実際に"
        "壊れるかは個別に確認してください。"
    )
    ws1["B3"].font = Font(name=FONT_NAME, size=9, italic=True, color="595959")
    ws1["B3"].alignment = Alignment(wrap_text=True)
    ws1.merge_cells("B3:H3")
    ws1.row_dimensions[3].height = 105

    legend_headers = ["Kind", "説明"]
    legend_rows = [
        ("Function", "namespaceスコープの関数(inline定義・プロトタイプ宣言とも)"),
        ("Variable", "inline/constexpr/externなnamespaceスコープの変数・配列定数"),
        ("Class", "class定義"),
        ("Struct", "struct定義"),
        ("Enum", "enum / enum class定義"),
        ("Using", "using エイリアス定義"),
    ]
    hrow = 5
    for i, h in enumerate(legend_headers):
        c = ws1[f"{get_column_letter(2 + i)}{hrow}"]
        c.value = h
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    for i, (kind, desc) in enumerate(legend_rows):
        row = hrow + 1 + i
        ws1[f"B{row}"] = kind
        ws1[f"C{row}"] = desc
        ws1[f"B{row}"].fill = KIND_FILL.get(kind, PatternFill())
        for col in "BC":
            cell = ws1[f"{col}{row}"]
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(col == "C"))
    ws1.column_dimensions["A"].width = 3
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 70

    # --- Sheet 2: findings -------------------------------------------------
    ws2 = wb.create_sheet("未インクルード疑い一覧")
    ws2.sheet_view.showGridLines = False
    headers = ["レイヤ", "使用元ファイル", "シンボル", "種別", "定義元ファイル", "推奨include文", "備考"]
    for i, h in enumerate(headers):
        c = ws2[f"{get_column_letter(i + 1)}1"]
        c.value = h
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    widths = [14, 42, 24, 10, 42, 46, 34]
    for i, w in enumerate(widths):
        ws2.column_dimensions[get_column_letter(i + 1)].width = w

    for ridx, r in enumerate(findings):
        row = ridx + 2
        values = [r.layer, r.file, r.symbol, r.kind, "; ".join(r.defined_in), r.suggested_include, r.note]
        for ci, v in enumerate(values):
            cell = ws2[f"{get_column_letter(ci + 1)}{row}"]
            cell.value = v
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=ci in (1, 4, 5, 6))
            if ci == 3:
                cell.fill = KIND_FILL.get(r.kind, PatternFill())

    ws2.freeze_panes = "C2"
    last_row = max(2, len(findings) + 1)
    ws2.auto_filter.ref = f"A1:G{last_row}"

    # --- Sheet 3: summary per file ------------------------------------------
    ws3 = wb.create_sheet("サマリー(ファイル別)")
    ws3.sheet_view.showGridLines = False
    ws3["B2"] = "ファイル別 該当件数"
    ws3["B2"].font = Font(name=FONT_NAME, size=14, bold=True, color="203864")
    headers3 = ["使用元ファイル", "該当件数"]
    hrow3 = 4
    for i, h in enumerate(headers3):
        c = ws3[f"{get_column_letter(2 + i)}{hrow3}"]
        c.value = h
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

    per_file: dict[str, int] = {}
    for r in findings:
        per_file[r.file] = per_file.get(r.file, 0) + 1
    findings_last_row = max(2, len(findings) + 1)
    ranked = sorted(per_file.items(), key=lambda kv: kv[1], reverse=True)
    for i, (fname, _n) in enumerate(ranked):
        row = hrow3 + 1 + i
        ws3[f"B{row}"] = fname
        ws3[f"C{row}"] = f"=COUNTIF('未インクルード疑い一覧'!$B$2:$B${findings_last_row},B{row})"
        for col in "BC":
            cell = ws3[f"{col}{row}"]
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=9)
            cell.alignment = Alignment(horizontal="left" if col == "B" else "center")
    ws3.column_dimensions["A"].width = 3
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["C"].width = 14

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def print_summary(findings: list[Finding]) -> None:
    print(f"# 未インクルード疑いの検出件数: {len(findings)}")
    if not findings:
        return
    per_file: dict[str, int] = {}
    for r in findings:
        per_file[r.file] = per_file.get(r.file, 0) + 1
    print()
    print("# ファイル別件数 Top15")
    for fname, n in sorted(per_file.items(), key=lambda kv: kv[1], reverse=True)[:15]:
        print(f"  - {fname}: {n}件")
    print()
    print("# 検出内容(先頭20件)")
    for r in findings[:20]:
        print(f"  - [{r.layer}] {r.file}: `{r.symbol}` ({r.kind}) は {', '.join(r.defined_in)} で定義 "
              f"-> 提案: {r.suggested_include}")
    if len(findings) > 20:
        print(f"  ... 他 {len(findings) - 20} 件 (CSV/xlsx出力を参照)")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--target", type=Path, default=None,
                         help="Folder to scan recursively (default: missing_include_audit_target/ next to this script)")
    parser.add_argument("--output", type=Path, default=None,
                         help="Folder to write results into (default: missing_include_audit_output/ next to this script)")
    parser.add_argument("--csv", type=Path, default=None, help="Override CSV output path (default: <output>/result.csv)")
    parser.add_argument("--xlsx", type=Path, default=None, help="Override Excel output path (default: <output>/result.xlsx, needs openpyxl)")
    args = parser.parse_args()

    target_dir = args.target.resolve() if args.target else target_dir_for(__file__)
    output_dir = args.output.resolve() if args.output else output_dir_for(__file__)

    if not target_dir.is_dir():
        print(f"Target folder not found: {target_dir}", file=sys.stderr)
        return 1

    print(f"# 解析対象フォルダ: {target_dir}")
    findings = scan_target(target_dir)
    findings.sort(key=lambda r: (r.file, r.symbol))

    print_summary(findings)

    csv_path = args.csv if args.csv else output_dir / "result.csv"
    write_csv(findings, csv_path)
    print(f"\nCSV written: {csv_path}")

    xlsx_path = args.xlsx if args.xlsx else output_dir / "result.xlsx"
    try:
        write_xlsx(findings, xlsx_path)
        print(f"Excel written: {xlsx_path}")
    except ImportError:
        print("openpyxl is not installed; skipping Excel output "
              "(pip install openpyxl)", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
