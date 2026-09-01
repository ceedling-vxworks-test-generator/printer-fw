#!/usr/bin/env python3
"""coverage_audit.py — mechanical unit-test coverage survey.

Scans every file placed under coverage_audit_target/ (next to this
script), regardless of file extension. Each file is classified by its
*content*, not its location or extension:

  * a file containing GoogleTest TEST()/TEST_F() macros is treated as
    a test file, and its test cases are collected;
  * anything else is treated as a source file, and its class'
    public methods plus any free (namespace-scope) functions are
    extracted.

For every extracted method/function, this script then reports whether
its name is referenced from inside a collected test body, and
classifies each referencing test case as "normal" or "abnormal" by
matching its test-case name against a keyword list (see
ABNORMAL_KEYWORDS below).

This is a *naming-convention heuristic*, not a semantic read of what a
test actually asserts. Two things it deliberately does NOT do:

  * It cannot tell two overloads of the same method apart (e.g.
    two `FindDataItem` overloads on the same class) — a method name
    match against a test file counts for every overload sharing that
    name. Rows for such methods are flagged in the "Note" column.

  * It does not judge *which* abnormal scenarios are missing, or
    propose new ones — it only reports what is/isn't exercised today,
    based on existing test names.

Usage:
    python3 tools/coverage_audit.py

    Copy the files you want analyzed (any format, any subfolder
    structure) into coverage_audit_target/, then run the script. It
    writes coverage_audit_output/result.csv and result.xlsx (needs
    openpyxl for the latter) and prints a summary to stdout.

    --target / --output / --csv / --xlsx let you override the default
    folders for one-off runs; see `--help`.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from _io_layout import iter_all_files, output_dir_for, target_dir_for  # noqa: E402

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# Substrings (case-insensitive) that, when found in a GoogleTest case
# name (Suite.Case, checked against "Case" — the part after the last
# dot survives typical renames better than the full string), mark that
# test as covering an abnormal / edge-case / negative path rather than
# the normal/happy path. Extend this list to match this repo's naming
# conventions as they evolve; it is intentionally conservative to avoid
# false positives (e.g. "False" is excluded because tests like
# PrintReadyBecomesFalseWhenDoorOpen describe a *normal* scenario).
ABNORMAL_KEYWORDS = [
    "null", "nullptr", "invalid", "unknown", "notfound", "not_found",
    "reject", "mismatch", "missing", "duplicate", "unregistered",
    "outofrange", "out_of_range", "timeout", "overflow", "returnfalse",
    "returnsfalse", "returnserror", "throws", "crash", "fail",
    "denied", "unavailable", "notinitialized", "uninitialized",
    "corrupt", "exceeds", "toolarge", "zerosize", "emptylist",
    "unresolvable", "unregistereddataid", "unregisteredid",
]

# Directory-path fragment -> display layer name. Longer/more specific
# fragments must come before shorter ones since the first match wins.
# Matched against the file's path *relative to the target folder*, so
# this only resolves to something other than "Other" when the folder
# structure copied into coverage_audit_target/ preserves the repo's
# original subpaths (e.g. core/adapter/... , products/printer_a/...).
LAYER_RULES = [
    ("core/storage/accessor", "Accessor"),
    ("core/storage/snapshot", "Snapshot"),
    ("core/storage", "Storage"),
    ("core/adapter", "Adapter"),
    ("core/capability", "Capability"),
    ("core/publisher", "Publisher"),
    ("core/store", "Store"),
    ("core/common", "Common"),
    ("include/rim_api.h", "C-API"),
    ("printer_a.h", "C-API"),
    ("products/printer_a", "Product(printer_a)"),
]

# ---------------------------------------------------------------------------
# Small C++-ish text utilities (regex/brace based — not a real parser)
# ---------------------------------------------------------------------------

_LINE_COMMENT_RE = re.compile(r"//[^\n]*")
_BLOCK_COMMENT_RE = re.compile(r"/\*.*?\*/", re.DOTALL)
_STRING_RE = re.compile(r'"(?:\\.|[^"\\])*"')
_CHAR_RE = re.compile(r"'(?:\\.|[^'\\])*'")


_PREPROCESSOR_LINE_RE = re.compile(r"^[ \t]*#.*$", re.MULTILINE)


def strip_preprocessor_directives(text: str) -> str:
    """Blank out preprocessor directive lines (#pragma/#include/#ifdef/...).

    They carry no C++ declarations of interest and, left in place, can
    get glued as leading noise onto the next real statement whenever no
    semicolon happens to separate them from it (e.g. a header whose very
    first declaration follows straight after `#include`s with no
    intervening typedef)."""
    return _PREPROCESSOR_LINE_RE.sub(lambda m: " " * len(m.group(0)), text)


def strip_comments_and_literals(text: str) -> str:
    """Blank out comments and string/char literals (keeps line numbers
    stable-ish by replacing with spaces, not deleting)."""

    def blank(m: re.Match) -> str:
        s = m.group(0)
        return "".join("\n" if c == "\n" else " " for c in s)

    text = _BLOCK_COMMENT_RE.sub(blank, text)
    text = _LINE_COMMENT_RE.sub(blank, text)
    text = _STRING_RE.sub(blank, text)
    text = _CHAR_RE.sub(blank, text)
    return text


def find_matching_brace(text: str, open_pos: int) -> int:
    """Given text[open_pos] == '{', return the index of its matching '}'."""
    depth = 0
    for i in range(open_pos, len(text)):
        if text[i] == "{":
            depth += 1
        elif text[i] == "}":
            depth -= 1
            if depth == 0:
                return i
    return len(text) - 1


def find_matching_paren(text: str, open_pos: int) -> int:
    depth = 0
    for i in range(open_pos, len(text)):
        if text[i] == "(":
            depth += 1
        elif text[i] == ")":
            depth -= 1
            if depth == 0:
                return i
    return len(text) - 1


def split_statements(text: str) -> list[str]:
    """Split a brace/paren-balanced text blob into top-level statements.

    A statement ends at a top-level ';', or (for definitions with an
    inline body) at the '}' matching a top-level '{'.
    """
    statements = []
    i = 0
    n = len(text)
    start = 0
    paren_depth = 0
    brace_depth = 0
    while i < n:
        c = text[i]
        if c == "(":
            paren_depth += 1
        elif c == ")":
            paren_depth = max(0, paren_depth - 1)
        elif c == "{" and paren_depth == 0 and brace_depth == 0:
            close = find_matching_brace(text, i)
            statements.append(text[start:close + 1])
            i = close + 1
            start = i
            continue
        elif c == "{":
            brace_depth += 1
        elif c == "}":
            brace_depth = max(0, brace_depth - 1)
        elif c == ";" and paren_depth == 0 and brace_depth == 0:
            statements.append(text[start:i + 1])
            i += 1
            start = i
            continue
        i += 1
    tail = text[start:].strip()
    if tail:
        statements.append(tail)
    return [s for s in statements if s.strip()]


_ACCESS_RE = re.compile(r"^\s*(public|private|protected)\s*:\s*;?\s*$")
_ACCESS_LABEL_RE = re.compile(r"\b(public|private|protected)\s*:")


def isolate_access_labels(body: str) -> str:
    """Force `public:` / `private:` / `protected:` labels to become their
    own statement when fed to split_statements(), by fencing them with
    synthetic semicolons. Without this, a label has no trailing ';' or
    '{' of its own and gets glued onto the following declaration."""
    return _ACCESS_LABEL_RE.sub(lambda m: f";{m.group(1)}:;", body)


def split_by_access(statements: list[str], default_access: str) -> list[tuple[str, str]]:
    """Tag each statement with the access level in force when it appears."""
    access = default_access
    out = []
    for s in statements:
        m = _ACCESS_RE.match(s.strip())
        if m:
            access = m.group(1)
            continue
        out.append((access, s))
    return out


_CLASS_RE = re.compile(
    r"(template\s*<[^;{}]*>\s*)?\b(class|struct)\s+(\w+)\s*(final\s*)?"
    r"(:\s*[^{};]*)?\{"
)


@dataclass
class MethodInfo:
    class_name: str
    access: str
    signature: str
    name: str
    overload_count: int = 1


_SKIP_LEADING_RE = re.compile(
    r"^\s*(template\s*<[^;{}]*>\s*)?"
    r"((virtual|static|inline|explicit|friend|constexpr|const)\s+)*"
)
_NAME_PAREN_RE = re.compile(r"([~]?\w+)\s*\(")
_CONTROL_KEYWORDS = {
    "if", "for", "while", "switch", "catch", "sizeof", "return",
    "new", "delete", "decltype", "static_cast", "dynamic_cast",
    "const_cast", "reinterpret_cast", "using",
}


def split_header_and_body(stmt: str) -> tuple[str, bool]:
    """Return (header_text, had_inline_body)."""
    stmt = stmt.rstrip()
    if stmt.endswith(";"):
        return stmt[:-1], False
    if stmt.endswith("}"):
        # find first top-level '{' that starts the body
        depth = 0
        for i, c in enumerate(stmt):
            if c == "(":
                depth += 1
            elif c == ")":
                depth = max(0, depth - 1)
            elif c == "{" and depth == 0:
                return stmt[:i], True
        return stmt, True
    return stmt, False


def parse_method_statement(stmt: str, class_name: str) -> MethodInfo | None:
    header, _has_body = split_header_and_body(stmt)
    header = header.strip()
    if not header or "(" not in header:
        return None

    matches = list(_NAME_PAREN_RE.finditer(header))
    if not matches:
        return None
    # The *first* "identifier(" in a well-formed declaration is the
    # function/constructor name itself; later matches are calls inside
    # default-argument expressions or (for constructors) the member
    # initializer list, e.g. "Foo(int x) : a_(x), b_(Bar())".
    name_m = matches[0]
    name = name_m.group(1)
    if name.lower() in _CONTROL_KEYWORDS:
        return None

    open_paren = header.index("(", name_m.start())
    close_paren = find_matching_paren(header, open_paren)
    if close_paren <= open_paren:
        return None
    trailer = header[close_paren + 1:].strip()

    # Drop a constructor's member-initializer list ("... ) : a_(x), b_(y)")
    # from the stored signature -- keep only up to the matched ')', plus
    # any qualifiers between ')' and ':' (const/override/noexcept/etc).
    colon_pos = trailer.find(":")
    if colon_pos != -1 and not trailer[:colon_pos].strip().startswith("::"):
        header_for_display = header[:close_paren + 1] + " " + trailer[:colon_pos].strip()
    else:
        header_for_display = header

    signature = re.sub(r"\s+", " ", header_for_display).strip()
    return MethodInfo(
        class_name=class_name,
        access="public",
        signature=signature,
        name=name,
    )


# Matches `namespace X {` / `extern "C" {`. Note: by the time this runs,
# strip_comments_and_literals() has already blanked string literals to
# spaces, so the literal `"C"` in `extern "C" {` has become whitespace --
# match `extern` followed by arbitrary whitespace then `{`, not the quotes.
_WRAPPER_RE = re.compile(r'(namespace\s+\w*\s*\{|extern\s*\{)')


def flatten_wrappers(text: str) -> str:
    """Repeatedly strip one layer of `namespace X { ... }` / `extern "C"
    { ... }` wrapper tokens so their contents become top-level statements,
    without discarding anything else in the wrapped body."""
    changed = True
    while changed:
        changed = False
        m = _WRAPPER_RE.search(text)
        if not m:
            break
        open_brace = m.end() - 1
        close_brace = find_matching_brace(text, open_brace)
        text = (
            text[:m.start()]
            + " " * len(m.group(0))
            + text[open_brace + 1:close_brace]
            + " "
            + text[close_brace + 1:]
        )
        changed = True
    return text


# ---------------------------------------------------------------------------
# Test-side scanning
# ---------------------------------------------------------------------------

_TEST_MACRO_RE = re.compile(
    r"\bTEST(?:_F|_P)?\s*\(\s*([\w:]+)\s*,\s*([\w]+)\s*\)\s*\{"
)


@dataclass
class TestCase:
    suite: str
    case: str
    file: Path
    body: str

    @property
    def full_name(self) -> str:
        return f"{self.suite}.{self.case}"

    def is_abnormal(self) -> bool:
        haystack = self.case.lower()
        return any(kw in haystack for kw in ABNORMAL_KEYWORDS)


def collect_test_cases(test_files: list[Path]) -> list[TestCase]:
    cases: list[TestCase] = []
    for path in test_files:
        raw = path.read_text(encoding="utf-8", errors="replace")
        text = strip_preprocessor_directives(strip_comments_and_literals(raw))
        for m in _TEST_MACRO_RE.finditer(text):
            open_brace = m.end() - 1
            close_brace = find_matching_brace(text, open_brace)
            body = text[open_brace + 1:close_brace]
            cases.append(TestCase(suite=m.group(1), case=m.group(2), file=path, body=body))
    return cases


_IDENT_BOUNDARY = r"(?<![\w])"
_IDENT_BOUNDARY_END = r"(?![\w])"


def name_referenced(name: str, text: str) -> bool:
    if not name:
        return False
    pattern = _IDENT_BOUNDARY + re.escape(name) + _IDENT_BOUNDARY_END
    return re.search(pattern, text) is not None


# ---------------------------------------------------------------------------
# Layer classification
# ---------------------------------------------------------------------------

def classify_layer(rel_path: Path) -> str:
    rel = rel_path.as_posix()
    for fragment, label in LAYER_RULES:
        if fragment in rel:
            return label
    return "Other"


# ---------------------------------------------------------------------------
# Row assembly
# ---------------------------------------------------------------------------

@dataclass
class CoverageRow:
    layer: str
    class_name: str
    method: str
    source_file: str
    name: str
    overload_count: int = 1
    normal_tested: bool = False
    normal_test_names: list[str] = field(default_factory=list)
    abnormal_tested: bool = False
    abnormal_test_names: list[str] = field(default_factory=list)

    @property
    def note(self) -> str:
        return "overloaded name (ambiguous match)" if self.overload_count > 1 else ""


def scan_source_file(path: Path, target_dir: Path) -> list[CoverageRow]:
    """Extract public class methods and free (namespace-scope) functions
    from a single non-test file, classifying by content only."""
    raw = path.read_text(encoding="utf-8", errors="replace")
    text = strip_preprocessor_directives(strip_comments_and_literals(raw))
    rel = path.relative_to(target_dir)
    layer = classify_layer(rel)

    rows: list[CoverageRow] = []
    class_spans: list[tuple[int, int]] = []

    for m in _CLASS_RE.finditer(text):
        kind = m.group(2)
        cname = m.group(3)
        open_brace = m.end() - 1
        close_brace = find_matching_brace(text, open_brace)
        class_spans.append((m.start(), close_brace))
        body = isolate_access_labels(text[open_brace + 1:close_brace])
        default_access = "public" if kind == "struct" else "private"
        for access, stmt in split_by_access(split_statements(body), default_access):
            if access != "public":
                continue
            info = parse_method_statement(stmt, cname)
            if info is not None:
                rows.append(CoverageRow(
                    layer=layer, class_name=info.class_name, method=info.signature,
                    source_file=str(rel), name=info.name,
                ))

    # Free functions live outside every class body -- carve those spans
    # out first so a method statement never gets double-counted here.
    outside_chunks = []
    cursor = 0
    for start, end in sorted(class_spans):
        if start > cursor:
            outside_chunks.append(text[cursor:start])
        cursor = max(cursor, end + 1)
    outside_chunks.append(text[cursor:])
    outside_text = flatten_wrappers("\n".join(outside_chunks))

    for stmt in split_statements(outside_text):
        info = parse_method_statement(stmt, path.name)
        if info is not None and info.name not in ("namespace",):
            rows.append(CoverageRow(
                layer=layer, class_name=info.class_name, method=info.signature,
                source_file=str(rel), name=info.name,
            ))

    return rows


def build_rows_from_target(target_dir: Path) -> list[CoverageRow]:
    test_files: list[Path] = []
    source_files: list[Path] = []
    for path in iter_all_files(target_dir):
        raw = path.read_text(encoding="utf-8", errors="replace")
        stripped = strip_preprocessor_directives(strip_comments_and_literals(raw))
        if _TEST_MACRO_RE.search(stripped):
            test_files.append(path)
        else:
            source_files.append(path)

    rows: list[CoverageRow] = []
    for path in source_files:
        rows.extend(scan_source_file(path, target_dir))

    # Overload counts are computed globally (class_name, name) so that a
    # method split across files (or genuinely overloaded within one) is
    # still flagged as an ambiguous match.
    counts: dict[tuple[str, str], int] = {}
    for r in rows:
        counts[(r.class_name, r.name)] = counts.get((r.class_name, r.name), 0) + 1
    for r in rows:
        r.overload_count = counts[(r.class_name, r.name)]

    test_cases = collect_test_cases(test_files)
    for row in rows:
        for tc in test_cases:
            if name_referenced(row.name, tc.body):
                if tc.is_abnormal():
                    row.abnormal_tested = True
                    row.abnormal_test_names.append(tc.full_name)
                else:
                    row.normal_tested = True
                    row.normal_test_names.append(tc.full_name)

    return rows


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def write_csv(rows: list[CoverageRow], out_path: Path) -> None:
    import csv
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "Layer", "Class", "Method", "NormalTested", "NormalTestNames",
            "AbnormalTested", "AbnormalTestNames", "SourceFile", "Note",
        ])
        for r in rows:
            w.writerow([
                r.layer, r.class_name, r.method,
                "YES" if r.normal_tested else "NO",
                "; ".join(sorted(set(r.normal_test_names))),
                "YES" if r.abnormal_tested else "NO",
                "; ".join(sorted(set(r.abnormal_test_names))),
                r.source_file, r.note,
            ])


def write_xlsx(rows: list[CoverageRow], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FONT_NAME = "Arial"
    HEADER_FILL = PatternFill("solid", fgColor="203864")
    HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    OK_FILL = PatternFill("solid", fgColor="C6E0B4")
    NG_FILL = PatternFill("solid", fgColor="F8CBAD")
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = Workbook()
    ws = wb.active
    ws.title = "カバレッジ一覧"
    headers = ["レイヤ", "クラス", "メソッド", "正常系", "正常系テスト",
               "異常系", "異常系テスト", "ソースファイル", "備考"]
    for i, h in enumerate(headers):
        c = ws[f"{get_column_letter(i + 1)}1"]
        c.value = h
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    widths = [16, 26, 50, 8, 40, 8, 40, 34, 24]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    for ridx, r in enumerate(rows):
        row = ridx + 2
        n_sym = "⭕" if r.normal_tested else "❌"
        a_sym = "⭕" if r.abnormal_tested else "❌"
        values = [
            r.layer, r.class_name, r.method, n_sym,
            "; ".join(sorted(set(r.normal_test_names))),
            a_sym, "; ".join(sorted(set(r.abnormal_test_names))),
            r.source_file, r.note,
        ]
        for ci, v in enumerate(values):
            cell = ws[f"{get_column_letter(ci + 1)}{row}"]
            cell.value = v
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=ci in (2, 4, 6))
            cell.font = Font(name=FONT_NAME, size=9)
            if ci == 3:
                cell.fill = OK_FILL if r.normal_tested else NG_FILL
            if ci == 5:
                cell.fill = OK_FILL if r.abnormal_tested else NG_FILL

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:I{len(rows) + 1}"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def print_summary(rows: list[CoverageRow]) -> None:
    total = len(rows)
    normal_ok = sum(1 for r in rows if r.normal_tested)
    abnormal_ok = sum(1 for r in rows if r.abnormal_tested)

    print(f"# 調査対象メソッド数: {total}")
    print(f"# 正常系テストあり: {normal_ok} / {total} "
          f"({0 if total == 0 else normal_ok / total:.1%})")
    print(f"# 異常系テストあり: {abnormal_ok} / {total} "
          f"({0 if total == 0 else abnormal_ok / total:.1%})")
    print()
    print(f"{'レイヤ':<20} {'件数':>6} {'正常系OK':>10} {'異常系OK':>10}")
    by_layer: dict[str, list[CoverageRow]] = {}
    for r in rows:
        by_layer.setdefault(r.layer, []).append(r)
    for layer, layer_rows in sorted(by_layer.items()):
        n = len(layer_rows)
        n_ok = sum(1 for r in layer_rows if r.normal_tested)
        a_ok = sum(1 for r in layer_rows if r.abnormal_tested)
        print(f"{layer:<20} {n:>6} {n_ok:>10} {a_ok:>10}")

    untested = [r for r in rows if not r.normal_tested and not r.abnormal_tested]
    if untested:
        print()
        print(f"# 正常系・異常系ともに未テストのメソッド: {len(untested)} 件")
        for r in untested[:30]:
            print(f"  - [{r.layer}] {r.class_name}::{r.method}")
        if len(untested) > 30:
            print(f"  ... 他 {len(untested) - 30} 件 (CSV/xlsx出力を参照)")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--target", type=Path, default=None,
                         help="Folder to scan recursively (default: coverage_audit_target/ next to this script)")
    parser.add_argument("--output", type=Path, default=None,
                         help="Folder to write results into (default: coverage_audit_output/ next to this script)")
    parser.add_argument("--csv", type=Path, default=None, help="Override CSV output path (default: <output>/result.csv)")
    parser.add_argument("--xlsx", type=Path, default=None, help="Override Excel output path (default: <output>/result.xlsx, needs openpyxl)")
    args = parser.parse_args()

    target_dir = args.target.resolve() if args.target else target_dir_for(__file__)
    output_dir = args.output.resolve() if args.output else output_dir_for(__file__)

    if not target_dir.is_dir():
        print(f"Target folder not found: {target_dir}", file=sys.stderr)
        return 1

    print(f"# 解析対象フォルダ: {target_dir}")
    rows = build_rows_from_target(target_dir)
    rows.sort(key=lambda r: (r.layer, r.class_name, r.method))

    print_summary(rows)

    csv_path = args.csv if args.csv else output_dir / "result.csv"
    write_csv(rows, csv_path)
    print(f"\nCSV written: {csv_path}")

    xlsx_path = args.xlsx if args.xlsx else output_dir / "result.xlsx"
    try:
        write_xlsx(rows, xlsx_path)
        print(f"Excel written: {xlsx_path}")
    except ImportError:
        print("openpyxl is not installed; skipping Excel output "
              "(pip install openpyxl)", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
